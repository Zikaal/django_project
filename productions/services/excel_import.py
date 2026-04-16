from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from productions.forms import DailyProductionForm
from productions.models import Well

# Поддерживаемые варианты названий колонок.
# Это делает импорт менее хрупким к разным Excel-шаблонам.
HEADER_ALIASES = {
    "well": {"well", "скважина", "название скважины", "well_name"},
    "date": {"date", "дата"},
    "work_time": {"work_time", "время работы", "часы", "hours"},
    "liquid_debit": {"liquid_debit", "дебит жидкости", "жидкость", "liquid"},
    "water_cut": {"water_cut", "обводненность", "water", "water_percent"},
    "oil_density": {"oil_density", "плотность нефти", "density"},
}


def normalize_header(value):
    """
    Приводит заголовок Excel-колонки к нормализованному виду:
    - строка;
    - lower();
    - без лишних пробелов и переводов строки.
    """
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ")


def parse_date_value(value):
    """
    Преобразует значение ячейки в date.

    Поддерживает:
    - datetime;
    - date;
    - строки нескольких форматов.
    """
    if value is None or value == "":
        raise ValueError("Дата не указана.")

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    value = str(value).strip()

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Не удалось распознать дату: {value}")


def parse_decimal_value(value, field_name):
    """
    Безопасно преобразует значение ячейки в Decimal.

    Поддерживает int/float/Decimal и строки.
    Запятую заменяет на точку.
    """
    if value is None or value == "":
        raise ValueError(f"Поле '{field_name}' не заполнено.")

    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    raw = str(value).strip().replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Поле '{field_name}' содержит некорректное число: {value}") from exc


def resolve_headers(header_row):
    """
    Находит индексы обязательных колонок по первой строке Excel.

    Возвращает:
    - словарь вида {"well": 0, "date": 1, ...}

    Если обязательная колонка не найдена — выбрасывает ValueError.
    """
    header_map = {}
    normalized_headers = [normalize_header(cell) for cell in header_row]

    for canonical_name, aliases in HEADER_ALIASES.items():
        found_index = None
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                found_index = index
                break
        if found_index is None:
            raise ValueError(f"Не найдена обязательная колонка: {canonical_name}")
        header_map[canonical_name] = found_index

    return header_map


def _import_from_workbook_source(file_source):
    """
    Внутренний сценарий импорта из workbook-источника.

    Логика:
    1. Открываем workbook.
    2. Проверяем, что файл не пустой.
    3. Разбираем заголовки.
    4. Идем по строкам.
    5. Пропускаем пустые строки.
    6. Ищем скважину по имени.
    7. Парсим дату и числа.
    8. Валидируем через DailyProductionForm.
    9. Сохраняем только корректные записи.
    """
    workbook = load_workbook(filename=file_source, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {
            "created_count": 0,
            "skipped_count": 0,
            "errors": ["Файл пустой."],
        }

    try:
        header_map = resolve_headers(rows[0])
    except ValueError as exc:
        return {
            "created_count": 0,
            "skipped_count": 0,
            "errors": [str(exc)],
        }

    created_count = 0
    skipped_count = 0
    errors = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(cell in (None, "") for cell in row):
            skipped_count += 1
            continue

        try:
            well_name_raw = row[header_map["well"]]
            well_name = str(well_name_raw).strip() if well_name_raw is not None else ""
            if not well_name:
                raise ValueError("Название скважины не указано.")

            well = Well.objects.get(name=well_name)

            parsed_date = parse_date_value(row[header_map["date"]])
            work_time = parse_decimal_value(row[header_map["work_time"]], "work_time")
            liquid_debit = parse_decimal_value(row[header_map["liquid_debit"]], "liquid_debit")
            water_cut = parse_decimal_value(row[header_map["water_cut"]], "water_cut")
            oil_density = parse_decimal_value(row[header_map["oil_density"]], "oil_density")

            # Используем существующую Django-форму как единый слой валидации.
            form = DailyProductionForm(
                data={
                    "well": well.pk,
                    "date": parsed_date.isoformat(),
                    "work_time": str(work_time),
                    "liquid_debit": str(liquid_debit),
                    "water_cut": str(water_cut),
                    "oil_density": str(oil_density),
                }
            )

            if form.is_valid():
                form.save()
                created_count += 1
            else:
                errors.append(f"Строка {row_number}: {form.errors.as_json()}")

        except Well.DoesNotExist:
            errors.append(f"Строка {row_number}: скважина '{well_name}' не найдена.")
        except Exception as exc:
            errors.append(f"Строка {row_number}: {str(exc)}")

    return {
        "created_count": created_count,
        "skipped_count": skipped_count,
        "errors": errors,
    }


def process_daily_productions_excel(file_path):
    """
    Публичная функция, которую вызывает Celery-задача из tasks.py.

    Принимает путь к Excel-файлу и передает его
    во внутренний обработчик импорта.
    """
    return _import_from_workbook_source(file_path)
