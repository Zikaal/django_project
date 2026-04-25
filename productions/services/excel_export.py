from decimal import Decimal
from io import BytesIO

from django.db.models import Avg, Count, DecimalField, ExpressionWrapper, F, Sum, Value
from django.db.models.functions import Coalesce
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from productions.models import DailyProduction


def build_monthly_production_report(year: int, month: int):
    """
    Строит Excel-отчет по рапортам за конкретный месяц.

    Что попадает в отчет:
    - компания;
    - скважина;
    - количество рапортов;
    - суммарный дебит жидкости;
    - средняя обводненность;
    - суммарная чистая нефть.

    Возвращает:
    - BytesIO с готовым .xlsx-файлом.
    """

    # Формула чистой нефти на уровне ORM:
    # liquid_debit * (1 - water_cut / 100) * oil_density
    oil_formula = ExpressionWrapper(
        F("liquid_debit") * (Value(Decimal("1.0")) - F("water_cut") / Value(Decimal("100.0"))) * F("oil_density"),
        output_field=DecimalField(max_digits=14, decimal_places=4),
    )

    queryset = (
        DailyProduction.objects.select_related("well", "well__oil_company")
        .filter(date__year=year, date__month=month)
        .values("well__oil_company__name", "well__name")
        .annotate(
            reports_count=Count("id"),
            total_liquid_debit=Coalesce(
                Sum("liquid_debit"),
                0,
                output_field=DecimalField(max_digits=14, decimal_places=2),
            ),
            avg_water_cut=Coalesce(
                Avg("water_cut"),
                0,
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
            total_oil=Coalesce(
                Sum(oil_formula),
                0,
                output_field=DecimalField(max_digits=14, decimal_places=4),
            ),
        )
        .order_by("well__oil_company__name", "well__name")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    # Заголовок листа.
    title = f"Сводный отчёт по скважинам за {month:02d}.{year}"
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Заголовки таблицы.
    headers = [
        "Компания",
        "Скважина",
        "Количество рапортов",
        "Суммарный дебит жидкости",
        "Средняя обводненность",
        "Суммарная чистая нефть",
    ]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_num = 4
    total_reports_sum = 0
    total_liquid_sum = Decimal("0")
    total_oil_sum = Decimal("0")

    # Заполняем строки отчета и параллельно считаем итоги.
    for item in queryset:
        ws.cell(row=row_num, column=1, value=item["well__oil_company__name"])
        ws.cell(row=row_num, column=2, value=item["well__name"])
        ws.cell(row=row_num, column=3, value=item["reports_count"])
        ws.cell(row=row_num, column=4, value=float(item["total_liquid_debit"]))
        ws.cell(row=row_num, column=5, value=float(item["avg_water_cut"]))
        ws.cell(row=row_num, column=6, value=float(item["total_oil"]))

        total_reports_sum += item["reports_count"]
        total_liquid_sum += item["total_liquid_debit"]
        total_oil_sum += item["total_oil"]

        row_num += 1

    # Итоговая строка.
    ws.cell(row=row_num, column=1, value="ИТОГО")
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=3, value=total_reports_sum)
    ws.cell(row=row_num, column=4, value=float(total_liquid_sum))
    ws.cell(row=row_num, column=6, value=float(total_oil_sum))

    # Дальше обычно идут автоширины колонок и сохранение в BytesIO.
    # Логика файла уже ориентирована на отдачу в Celery-задачу экспорта.
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
